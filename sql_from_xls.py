import openpyxl

# This script helps to make .sql file with INSERT statements for offliner's devices from .xlsx file

xlsx_file_path = '/Users/konstantinprilepin/work/offliner/импорт из xls/offliner_devices_from_13.03.2022_1709_.xlsx'
locations_sql_file_path = '/Users/konstantinprilepin/work/offliner/импорт из xls/locations.sql'
devices_sql_file_path = '/Users/konstantinprilepin/work/offliner/импорт из xls/devices.sql'


def fix_type_of_display(row):
    val = str(row[5].value).lower()
    mapping = {'Ситиформат': 2, 'Суперсайт': 5, 'Билборд': 4, 'Гологра': 1, 'Медиафасад': 11, 'Ситиборд': 3}
    for k, v in mapping.items():
        if k.lower() in val:
            return v
    return 2


field_to_col_map_location = {
    'id': 1,
    'placement': lambda row: 1 if row[7].value == 'помещение' else 2,
    'lat': 9,
    'lon': 10,
    'address': 8,
    'name': 8,
}

field_to_col_map_device = {
    'id': 1,
    'location_id': 1,
    'external_id': 2,
    'name': 3,
    'width': 4,
    'height': 5,
    'type': fix_type_of_display,
    'description': 11,
    'OTS': 12,
    'GRP': 13,
    'media_switch_delay': lambda row: '00:00:01',
}


def do(mapping, file_name, table_name):
    wb_obj = openpyxl.load_workbook(xlsx_file_path)

    sheet_obj = wb_obj.active
    with open(file_name, 'wt') as f:
        # for i in range(1, 5):
        for i in range(1, sheet_obj.max_row + 1):
            col_names = []
            values = []
            for col_name, getter in mapping.items():
                col_names.append(col_name)
                if isinstance(getter, int):
                    values.append("'" + str(sheet_obj.cell(row=i, column=getter).value) + "'")
                else:
                    values.append("'" + str(getter(sheet_obj[i])) + "'")
            f.write(f'''INSERT INTO {table_name}({','.join(col_names)}) VALUES ({','.join(values)});\n''')


if __name__ == '__main__':
    do(
        field_to_col_map_location,
        locations_sql_file_path,
        'offliner_location',
    )
    do(
        field_to_col_map_device,
        devices_sql_file_path,
        'offliner_device',
    )
